# -*- coding: utf-8 -*-
"""
Created on Thu Mar  8 13:10:14 2018

@author: Se

实现信息的汇总
"""
import os
import re
import openpyxl

def unique_index(L,e):
    return [i for (i,j) in enumerate(L) if j == e]

def find_in_sheet(sheet):
    row_num = []
    col_num = []
    row_temp = 0
    for row in sheet.iter_rows(min_row = 1, max_col = sheet.max_column, max_row = sheet.max_row):
        col_temp = 0
        for cell in row:
            #print(cell.value)
            temp = cell.value
            string = str(temp)
            string = re.sub(' ', '', string)
            string = re.sub(' ', '', string)
            if temp != None and '项目' in string:
                #print('T')
                row_num.append(row_temp)
                col_num.append(col_temp)
            col_temp += 1
        row_temp += 1
    
    row_set = set(row_num)
    row_left = 99
    for item in row_set:
        temp = row_num.count(item)
        if temp == 2:
            row_left = item
            
    save = unique_index(row_num, row_left)        
    return row_left, [col_num[save[0]], col_num[save[1]]]
    

if __name__ == '__main__':
    
    path = os.listdir('.')
    #print (path)
    names = []
    palaces = []
    dates = []

    for i in range(len(path)):
        temp = path[i]
        date = temp.split('.')[0]
        types = temp.split('.')[1]
        if types == 'xlsx' and date != '江山汇总':
            wb = openpyxl.reader.excel.load_workbook(temp)
            sheetnames = wb.get_sheet_names()  
            ws = wb.get_sheet_by_name(sheetnames[0])        
            r_num, c_num = find_in_sheet(ws)
            if len(c_num) != 2:
                print(i)
                continue
            
            for rownum in range(r_num + 1, ws.max_row):  
                col_temp = 0
                for cell in list(ws.rows)[rownum]:
                    if col_temp == c_num[0]:
                        names.append(cell.value)
                    elif col_temp == c_num[1]:
                        palaces.append(cell.value)
                        dates.append(date)
                    col_temp += 1    
        else:
            print(temp + '\t不合法文件')
                
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '买房'
    ws_out['A1'] = '序号'
    ws_out['B1'] = '名称'
    ws_out['C1'] = '地点'
    ws_out['D1'] = '时间'
    index = 2
    
    for i, val in enumerate(names):
        temp = 'A' + str(index)
        ws_out[temp] = i + 1
        temp = 'B' + str(index)
        ws_out[temp] = val
        temp = 'C' + str(index)
        ws_out[temp] = palaces[i]
        temp = 'D' + str(index)
        ws_out[temp] = dates[i]
        index += 1    
    
    wb_out.save('江山汇总.xlsx') 


