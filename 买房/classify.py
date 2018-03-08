# -*- coding: utf-8 -*-
"""
Created on Thu Mar  8 17:13:37 2018

@author: Se
"""

import openpyxl

if __name__ == '__main__':
    
    wb = openpyxl.reader.excel.load_workbook('江山汇总.xlsx')
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[0])

    # 获取表头
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数 
    city = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        city.append(app)
    
    daoli = []
    nangang = []
    daowai = []
    xiangfang = []
    pingfang = []
    taiping = []
    songbei = []
    kaifa = []
    wuchang = []
    binxian = []
    tonghe = []
    yanshou = []
    binzhou = []
    yilan = []
    shuangcheng = []
    binxi = []
    other = []
    shangzhi = []
    dongli = []
    acheng = []
    hulan = []
    
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '地区'
    ws_out['A1'] = '地区'    
    
    index = 2
    for i in city:
        temp = i['地点']
        temp_save = 'A' + str(index)
        if '道里' in temp:
            daoli.append(i)
            ws_out[temp_save] = '道里'
        elif '南岗' in temp:
            nangang.append(i)
            ws_out[temp_save] = '南岗'
        elif '道外' in temp:
            daowai.append(i)
            ws_out[temp_save] = '道外'
        elif '香坊' in temp:
            xiangfang.append(i)
            ws_out[temp_save] = '香坊'
        elif '平房' in temp:
            pingfang.append(i)
            ws_out[temp_save] = '平房'
        elif '太平' in temp:
            taiping.append(i)
            ws_out[temp_save] = '太平'
        elif '松北' in temp:
            songbei.append(i)
            ws_out[temp_save] = '松北'
        elif '开发' in temp:
            kaifa.append(i)
            ws_out[temp_save] = '开发'
        elif '五常' in temp:
            wuchang.append(i)
            ws_out[temp_save] = '五常'
        elif '宾县' in temp:
            binxian.append(i)
            ws_out[temp_save] = '宾县'
        elif '通河' in temp:
            tonghe.append(i)
            ws_out[temp_save] = '通河'
        elif '延寿' in temp:
            yanshou.append(i)
            ws_out[temp_save] = '延寿'
        elif '宾州' in temp:
            binzhou.append(i)
            ws_out[temp_save] = '宾州'
        elif '依兰' in temp:
            yilan.append(i)
            ws_out[temp_save] = '依兰'
        elif '双城' in temp:
            shuangcheng.append(i)
            ws_out[temp_save] = '双城'
        elif '宾西' in temp:
            binxi.append(i)
            ws_out[temp_save] = '宾西'
        elif '尚志' in temp:
            shangzhi.append(i)
            ws_out[temp_save] = '尚志'
        elif '动力' in temp:
            dongli.append(i)
            ws_out[temp_save] = '动力'
        elif '阿城' in temp:
            acheng.append(i)
            ws_out[temp_save] = '阿城'
        elif '呼兰' in temp:
            hulan.append(i)
            ws_out[temp_save] = '呼兰'
        else:
            other.append(i)
            ws_out[temp_save] = '其它'
        index += 1
  
        wb_out.save('地区分类.xlsx')  
