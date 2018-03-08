# -*- coding: utf-8 -*-
"""
Created on Sun Feb 11 11:53:40 2018

@author: Se
"""

import openpyxl
from phone import Phone

if __name__ == '__main__':
    
    p  = Phone()
    wb = openpyxl.reader.excel.load_workbook('备选.xlsx')
    sheetnames = wb.get_sheet_names()  
    ws = wb.get_sheet_by_name(sheetnames[3])
    
    # 获取表头
    colnames = []
    for cell in list(ws.rows)[0]:
        colnames.append(cell.value)

    # 获取所有数据
    nrows = ws.max_row #行数 
    nums = []  
    for rownum in range(1, nrows):  
        app = {} 
        index = 0
        for cell in list(ws.rows)[rownum]:
            app[colnames[index]] = cell.value
            index += 1
        nums.append(app)
    
    print('Read Done!')
    
    for i in range(1, len(nums)):
        temp = int(nums[i]['号码'])
        out = p.find(temp)
        if out != None:
            # 关于保存位置的修改在此处
            temp = 'C' + str((i + 2))
            ws[temp].value = out['province']
            temp = 'D' + str((i + 2))
            ws[temp].value = out['city']            
            print(out['province'] + out['city'])         
    wb.save('输出.xlsx')
